"""
Excel report generator with multiple sheets and charts.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelReportGenerator:
    """Generates professional Excel reports for ETL pipeline results."""
    
    def __init__(self, output_dir: str = "data/output/reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_pipeline_report(
        self,
        metrics: dict,
        data_sample: pd.DataFrame,
        pipeline_id: str
    ) -> Path:
        """
        Generate complete Excel report with multiple sheets.
        
        Args:
            metrics: Pipeline metrics dictionary
            data_sample: Sample of processed data
            pipeline_id: Pipeline execution ID
        
        Returns:
            Path to generated Excel file
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"etl_report_{pipeline_id}_{timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        wb = Workbook()
        
        self._create_summary_sheet(wb, metrics)
        self._create_data_sheet(wb, data_sample)
        self._create_metrics_sheet(wb, metrics)
        
        wb.save(filepath)
        
        return filepath
    
    def _create_summary_sheet(self, wb: Workbook, metrics: dict):
        """Create executive summary sheet."""
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        ws = wb.create_sheet('Resumen Ejecutivo', 0)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        ws['A1'] = 'REPORTE ETL - CONCILIACIÓN OPERATIVA'
        ws['A1'].font = Font(bold=True, size=16, color="366092")
        ws.merge_cells('A1:D1')
        
        ws['A3'] = 'Pipeline ID:'
        ws['B3'] = metrics.get('pipeline_id', 'N/A')
        
        ws['A4'] = 'Fecha Ejecución:'
        ws['B4'] = metrics.get('start_time', 'N/A')
        
        ws['A5'] = 'Estado:'
        ws['B5'] = metrics.get('status', 'N/A').upper()
        status_fill = PatternFill(
            start_color="00C851" if metrics.get('status') == 'success' else "FF4444",
            fill_type="solid"
        )
        ws['B5'].fill = status_fill
        ws['B5'].font = Font(bold=True, color="FFFFFF")
        
        ws['A7'] = 'MÉTRICAS CLAVE'
        ws['A7'].font = header_font
        ws['A7'].fill = header_fill
        ws.merge_cells('A7:D7')
        
        row = 8
        summary_data = [
            ('Total Registros Entrada', metrics.get('input_rows', 0)),
            ('Registros Validados', metrics.get('validation_passed', 0)),
            ('Registros Inválidos', metrics.get('validation_failed', 0)),
            ('Duplicados Removidos', metrics.get('duplicates_removed', 0)),
            ('Registros Cargados', metrics.get('rows_loaded', 0)),
            ('Tasa de Éxito', f"{metrics.get('rows_loaded', 0) / max(metrics.get('input_rows', 1), 1):.2%}"),
            ('Tiempo de Procesamiento', f"{metrics.get('processing_time_seconds', 0):.2f}s"),
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 30
    
    def _create_data_sheet(self, wb: Workbook, data_sample: pd.DataFrame):
        """Create data sample sheet."""
        ws = wb.create_sheet('Muestra de Datos')
        
        for r in dataframe_to_rows(data_sample.head(100), index=False, header=True):
            ws.append(r)
        
        header_fill = PatternFill(start_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_metrics_sheet(self, wb: Workbook, metrics: dict):
        """Create detailed metrics sheet with charts."""
        ws = wb.create_sheet('Métricas Detalladas')
        
        ws['A1'] = 'ANÁLISIS DE PROCESAMIENTO'
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = 'Distribución de Registros'
        
        data = [
            ['Categoría', 'Cantidad'],
            ['Validados', metrics.get('validation_passed', 0)],
            ['Inválidos', metrics.get('validation_failed', 0)],
            ['Duplicados', metrics.get('duplicates_removed', 0)],
            ['Cargados', metrics.get('rows_loaded', 0)],
        ]
        
        for row_idx, row_data in enumerate(data, 4):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        chart = PieChart()
        labels = Reference(ws, min_col=1, min_row=5, max_row=8)
        data_ref = Reference(ws, min_col=2, min_row=4, max_row=8)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = "Distribución de Registros"
        
        ws.add_chart(chart, "D4")


class SummaryReportGenerator:
    """Generates summary reports from metrics."""
    
    @staticmethod
    def generate_text_summary(metrics: dict) -> str:
        """Generate plain text summary."""
        success_rate = metrics.get('rows_loaded', 0) / max(metrics.get('input_rows', 1), 1)
        
        summary = f"""
========================================
REPORTE ETL - RESUMEN EJECUTIVO
========================================

Pipeline ID: {metrics.get('pipeline_id', 'N/A')}
Fecha: {metrics.get('start_time', 'N/A')}
Estado: {metrics.get('status', 'N/A').upper()}

MÉTRICAS:
- Total Entrada: {metrics.get('input_rows', 0):,} registros
- Validados: {metrics.get('validation_passed', 0):,} registros
- Inválidos: {metrics.get('validation_failed', 0):,} registros
- Duplicados Removidos: {metrics.get('duplicates_removed', 0):,}
- Cargados: {metrics.get('rows_loaded', 0):,} registros
- Tasa de Éxito: {success_rate:.2%}
- Tiempo: {metrics.get('processing_time_seconds', 0):.2f}s

========================================
"""
        return summary